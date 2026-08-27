"""Parse Unchained Labs Aunty Excel exports into long-form curves and a plate JSON.

`Analysis_graph` holds repeating per-well column blocks keyed on
`(file name, well)`. The block width depends on which series the export
included: a full thermal ramp is eight columns, a Tm/Tagg ramp without
Z-average is six, and isothermal runs are two or four. Exports that omit
the graph sheet still have `Analysis_table`, which is enough for a plate.
"""

from __future__ import annotations
import csv
import json
import logging
import math
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

FLAVOR_THERMAL = "thermal_ramp"
FLAVOR_SIZING = "sizing"
FLAVOR_ISOTHERMAL = "isothermal"

SHEET_GRAPH = "Analysis_graph"
SHEET_TABLE = "Analysis_table"
SHEET_INFO = "Experiment info"

# A sparkline cell is at most 64 screen pixels wide, so this is already about
# one point per two pixels. More would only inflate the plate JSON.
MAX_THUMBNAIL_POINTS = 32

# Excel hands back floats like 330.1234567890123. The viewer rounds every
# coordinate to 2 decimals when it draws, so the extra digits are dead weight.
THUMBNAIL_SIGNIFICANT_DIGITS = 6

CURVE_CSV_COLUMNS = [
    "file_name",
    "analysis_mode",
    "well",
    "sample",
    "series",
    "x",
    "y",
]

# Internal series ids written to CSV / JSON. Keep these stable; the web
# viewer keys display metadata off them.
SERIES_FLUORESCENCE = "fluorescence"
SERIES_DIFFERENTIAL = "differential"
SERIES_Z_AVG = "z_average_diameter"
SERIES_SLS = "sls"
SERIES_CORRELATION = "correlation"
SERIES_INTENSITY = "intensity"
SERIES_MASS = "mass"

# (x header, y header) → (flavor, series id). Sizing has two identical
# hydrodynamic_diameter/amplitude pairs; the first is intensity, the second
# mass — handled in `_series_pairs_from_headers`.
_SERIES_PAIR_KEYS: tuple[tuple[str, str, str, str], ...] = (
    ("temperature_fluorescence", "fluorescence", FLAVOR_THERMAL, SERIES_FLUORESCENCE),
    ("temperature_differential", "differential", FLAVOR_THERMAL, SERIES_DIFFERENTIAL),
    ("temperature_DLS", "Z-average diameter", FLAVOR_THERMAL, SERIES_Z_AVG),
    ("temperature_DLS", "SLS", FLAVOR_THERMAL, SERIES_SLS),
    ("time", "amplitude", FLAVOR_SIZING, SERIES_CORRELATION),
    ("time_fluorescence", "fluorescence", FLAVOR_ISOTHERMAL, SERIES_FLUORESCENCE),
    ("time_DLS", "SLS", FLAVOR_ISOTHERMAL, SERIES_SLS),
)

_SERIES_ROW_FIRST_CELLS = frozenset(
    {
        "temperature_fluorescence",
        "temperature_differential",
        "temperature_DLS",
        "time",
        "time_fluorescence",
        "time_DLS",
        "fluorescence",
        "hydrodynamic_diameter",
    }
)

_PRIMARY_SERIES = {
    FLAVOR_THERMAL: SERIES_FLUORESCENCE,
    FLAVOR_SIZING: SERIES_CORRELATION,
    FLAVOR_ISOTHERMAL: SERIES_FLUORESCENCE,
}

# Analysis_table columns → compact JSON keys. Tm of 0 means "no transition".
_TABLE_VALUE_KEYS: dict[str, str] = {
    "Tm1 (°C)": "tm1",
    "Tm2 (°C)": "tm2",
    "Tm3 (°C)": "tm3",
    "Tagg (°C)": "tagg",
    "Tonset (°C)": "tonset",
    "Tsize (°C)": "tsize",
    "Z-avg diameter (nm)": "z_avg_diameter",
    "PDI": "pdi",
    "Pk 1 diameter (nm)": "pk1_diameter",
    "Pk 1 intensity (%)": "pk1_intensity",
    "Pk 1 mass (%)": "pk1_mass",
    "fluorescence k₁ (s⁻¹)": "fluor_k1",
    "fluorescence k₂ (s⁻¹)": "fluor_k2",
    "fluorescence k₃ (s⁻¹)": "fluor_k3",
    "fluorescence R²": "fluor_r2",
    "SLS k₁ (s⁻¹)": "sls_k1",
    "SLS k₂ (s⁻¹)": "sls_k2",
    "SLS k₃ (s⁻¹)": "sls_k3",
    "SLS R²": "sls_r2",
}

_TM_VALUE_KEYS = frozenset({"tm1", "tm2", "tm3"})

# Experiment-info labels we keep as run metadata.
_INFO_METADATA_KEYS: dict[str, str] = {
    "Analysis mode": "analysis_mode",
    "Start temp (°C)": "start_temp_c",
    "End temp (°C)": "end_temp_c",
    "Rate (°C/min)": "rate_c_per_min",
    "Temperature (°C)": "temperature_c",
    "Analysis version": "analysis_version",
    "Experiment name": "experiment_name",
}


@dataclass(frozen=True)
class CurveRow:
    file_name: str
    analysis_mode: str | None
    well: str
    sample: str | None
    series: str
    x: float
    y: float


@dataclass
class WellBlock:
    file_name: str
    analysis_mode: str | None
    well: str
    sample: str | None
    flavor: str
    series: dict[str, list[list[float]]]


@dataclass
class TableWell:
    file_name: str
    well: str
    sample: str | None
    analysis_mode: str | None
    values: dict[str, float | None]


@dataclass
class ParsedAunty:
    metadata: dict[str, Any]
    experiments: list[dict[str, Any]]
    curve_rows: list[CurveRow]


class AuntyParseError(ValueError):
    """Raised when an Aunty workbook is missing expected sheets or headers."""


def parse_aunty_workbook(path: Path) -> ParsedAunty:
    """Read an Aunty `.xlsx` export into curve rows and a plate JSON payload."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        graph_rows = _sheet_rows(wb[SHEET_GRAPH]) if SHEET_GRAPH in wb.sheetnames else []
        table_rows = _sheet_rows(wb[SHEET_TABLE]) if SHEET_TABLE in wb.sheetnames else []
        info_rows = _sheet_rows(wb[SHEET_INFO]) if SHEET_INFO in wb.sheetnames else []
        has_graph = SHEET_GRAPH in wb.sheetnames
    finally:
        wb.close()

    blocks = _parse_graph_blocks(graph_rows) if graph_rows else []
    table_wells = _parse_analysis_table(table_rows)
    summaries = {(row.file_name, row.well): row.values for row in table_wells}

    if blocks:
        experiments = _build_experiments(blocks, summaries)
        curve_rows = _curve_rows_from_blocks(blocks)
    elif table_wells:
        # Some exports skip Analysis_graph and only write the summary table.
        experiments = _build_experiments_from_table(table_wells)
        curve_rows = []
    elif not has_graph:
        raise AuntyParseError(f"Workbook is missing the '{SHEET_GRAPH}' sheet")
    else:
        raise AuntyParseError("No well blocks found in Analysis_graph")

    metadata = _parse_experiment_info(info_rows, experiments)
    return ParsedAunty(metadata=metadata, experiments=experiments, curve_rows=curve_rows)


def write_curves_csv(path: Path, rows: list[CurveRow]) -> None:
    """Write long-form curve points for the well-detail chart."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=CURVE_CSV_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "file_name": row.file_name,
                    "analysis_mode": row.analysis_mode or "",
                    "well": row.well,
                    "sample": row.sample or "",
                    "series": row.series,
                    "x": row.x,
                    "y": row.y,
                }
            )


def write_plate_json(path: Path, experiments: list[dict[str, Any]]) -> None:
    """Write the downsampled plate payload used to render the sparkline grid."""
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"experiments": experiments}
    path.write_text(json.dumps(payload, allow_nan=False), encoding="utf-8")


def downsample_points(
    points: list[list[float]], max_points: int = MAX_THUMBNAIL_POINTS
) -> list[list[float]]:
    """Keep endpoints and evenly spaced interior points for sparkline thumbs."""
    n = len(points)
    if n <= max_points or max_points < 2:
        return points
    indices: list[int] = []
    seen: set[int] = set()
    for i in range(max_points):
        idx = round(i * (n - 1) / (max_points - 1))
        if idx not in seen:
            seen.add(idx)
            indices.append(idx)
    return [points[i] for i in indices]


def round_points(points: list[list[float]]) -> list[list[float]]:
    """Drop float noise from thumbnail points, keeping significant digits."""
    return [[_significant(x), _significant(y)] for x, y in points]


def _significant(value: float) -> float:
    # Significant digits, not decimal places: sizing correlation x values are
    # around 4e-07 and would round to zero.
    return float(f"{value:.{THUMBNAIL_SIGNIFICANT_DIGITS}g}")


def _sheet_rows(ws: Any) -> list[tuple[Any, ...]]:
    return [tuple(row) for row in ws.iter_rows(values_only=True)]


def _parse_graph_blocks(rows: list[tuple[Any, ...]]) -> list[WellBlock]:
    if not rows:
        return []

    file_row_i = _find_row(rows, "File name")
    well_row_i = _find_row(rows, "Well")
    sample_row_i = _find_row(rows, "Sample")
    if file_row_i is None or well_row_i is None:
        raise AuntyParseError("Analysis_graph is missing File name / Well header rows")

    mode_row_i = _find_row(rows, "Analysis mode")
    last_header_i = well_row_i if sample_row_i is None else max(well_row_i, sample_row_i)
    series_row_i = _find_series_row(rows, after=last_header_i)
    if series_row_i is None:
        raise AuntyParseError("Analysis_graph is missing a series-name row")

    starts = [i for i, value in enumerate(rows[file_row_i]) if value == "File name"]
    if not starts:
        return []
    width = starts[1] - starts[0] if len(starts) > 1 else len(rows[series_row_i]) - starts[0]
    if width <= 0:
        raise AuntyParseError("Could not determine Analysis_graph block width")

    data_rows = rows[series_row_i + 1 :]
    blocks: list[WellBlock] = []
    for start in starts:
        headers = _row_slice(rows[series_row_i], start, width)
        detected = _series_pairs_from_headers(headers)
        if detected is None:
            logger.warning("Skipping Analysis_graph block with unknown series %s", headers)
            continue
        flavor, pairs = detected
        file_name = _cell_str(_row_slice(rows[file_row_i], start, width)[1])
        well = _cell_str(_row_slice(rows[well_row_i], start, width)[1])
        if not (file_name and well):
            continue
        analysis_mode = None
        if mode_row_i is not None:
            analysis_mode = _cell_str(_row_slice(rows[mode_row_i], start, width)[1])
        sample = None
        if sample_row_i is not None:
            sample = _cell_str(_row_slice(rows[sample_row_i], start, width)[1])

        series: dict[str, list[list[float]]] = {}
        for series_id, x_offset, y_offset in pairs:
            points: list[list[float]] = []
            for data_row in data_rows:
                x = _num(_cell_at(data_row, start + x_offset))
                y = _num(_cell_at(data_row, start + y_offset))
                if x is None or y is None:
                    continue
                points.append([x, y])
            series[series_id] = points

        blocks.append(
            WellBlock(
                file_name=file_name,
                analysis_mode=analysis_mode or None,
                well=well,
                sample=sample or None,
                flavor=flavor,
                series=series,
            )
        )
    return blocks


def _parse_analysis_table(rows: list[tuple[Any, ...]]) -> list[TableWell]:
    if not rows:
        return []
    header = [_cell_str(v) or "" for v in rows[0]]
    try:
        file_i = header.index("File name")
        well_i = header.index("Well")
    except ValueError:
        logger.warning("Analysis_table is missing File name / Well columns")
        return []

    sample_i = header.index("Sample") if "Sample" in header else None
    mode_i = header.index("Analysis mode") if "Analysis mode" in header else None
    index_by_key = {name: header.index(name) for name in _TABLE_VALUE_KEYS if name in header}
    out: list[TableWell] = []
    for row in rows[1:]:
        file_name = _cell_str(_cell_at(row, file_i))
        well = _cell_str(_cell_at(row, well_i))
        if not (file_name and well):
            continue
        values: dict[str, float | None] = {}
        for header_name, key in _TABLE_VALUE_KEYS.items():
            col = index_by_key.get(header_name)
            if col is None:
                continue
            values[key] = _table_value(key, _num(_cell_at(row, col)))
        out.append(
            TableWell(
                file_name=file_name,
                well=well,
                sample=_cell_str(_cell_at(row, sample_i)) if sample_i is not None else None,
                analysis_mode=_cell_str(_cell_at(row, mode_i)) if mode_i is not None else None,
                values=values,
            )
        )
    return out


def _build_experiments(
    blocks: list[WellBlock],
    summaries: dict[tuple[str, str], dict[str, float | None]],
) -> list[dict[str, Any]]:
    grouped: dict[str, list[WellBlock]] = defaultdict(list)
    order: list[str] = []
    for block in blocks:
        if block.file_name not in grouped:
            order.append(block.file_name)
        grouped[block.file_name].append(block)

    experiments: list[dict[str, Any]] = []
    for file_name in order:
        well_blocks = grouped[file_name]
        flavor = well_blocks[0].flavor
        analysis_mode = next((b.analysis_mode for b in well_blocks if b.analysis_mode), None)
        wells: list[dict[str, Any]] = []
        for block in well_blocks:
            thumbs = {
                series_id: round_points(downsample_points(points))
                for series_id, points in block.series.items()
                if points
            }
            wells.append(
                {
                    "well": block.well,
                    "sample": block.sample,
                    "values": summaries.get((block.file_name, block.well), {}),
                    "series": thumbs,
                }
            )
        experiments.append(
            {
                "fileName": file_name,
                "analysisMode": analysis_mode,
                "flavor": flavor,
                "primarySeries": _PRIMARY_SERIES[flavor],
                "wells": wells,
            }
        )
    return experiments


def _build_experiments_from_table(table_wells: list[TableWell]) -> list[dict[str, Any]]:
    flavor = _flavor_from_table_wells(table_wells)
    if flavor is None:
        raise AuntyParseError("Could not determine experiment type from Analysis_table")

    grouped: dict[str, list[TableWell]] = defaultdict(list)
    order: list[str] = []
    for row in table_wells:
        if row.file_name not in grouped:
            order.append(row.file_name)
        grouped[row.file_name].append(row)

    experiments: list[dict[str, Any]] = []
    for file_name in order:
        wells_in_file = grouped[file_name]
        analysis_mode = next((w.analysis_mode for w in wells_in_file if w.analysis_mode), None)
        experiments.append(
            {
                "fileName": file_name,
                "analysisMode": analysis_mode,
                "flavor": flavor,
                "primarySeries": _PRIMARY_SERIES[flavor],
                "wells": [
                    {
                        "well": row.well,
                        "sample": row.sample,
                        "values": row.values,
                        "series": {},
                    }
                    for row in wells_in_file
                ],
            }
        )
    return experiments


def _flavor_from_table_wells(table_wells: list[TableWell]) -> str | None:
    keys: set[str] = set()
    for row in table_wells:
        keys.update(k for k, value in row.values.items() if value is not None)
        keys.update(row.values)
    if keys & {"fluor_k1", "fluor_k2", "fluor_k3", "sls_k1", "sls_k2", "sls_k3"}:
        return FLAVOR_ISOTHERMAL
    if keys & {"tm1", "tm2", "tm3", "tagg", "tonset", "tsize"}:
        return FLAVOR_THERMAL
    if keys & {"z_avg_diameter", "pdi", "pk1_diameter", "pk1_intensity", "pk1_mass"}:
        return FLAVOR_SIZING
    return None


def _curve_rows_from_blocks(blocks: list[WellBlock]) -> list[CurveRow]:
    rows: list[CurveRow] = []
    for block in blocks:
        for series_id, points in block.series.items():
            for x, y in points:
                rows.append(
                    CurveRow(
                        file_name=block.file_name,
                        analysis_mode=block.analysis_mode,
                        well=block.well,
                        sample=block.sample,
                        series=series_id,
                        x=x,
                        y=y,
                    )
                )
    return rows


def _parse_experiment_info(
    rows: list[tuple[Any, ...]],
    experiments: list[dict[str, Any]],
) -> dict[str, Any]:
    flavors = list(dict.fromkeys(exp["flavor"] for exp in experiments))
    metadata: dict[str, Any] = {
        "experiment_type": flavors[0] if len(flavors) == 1 else flavors,
        "experiment_count": len(experiments),
    }

    blocks = _info_blocks(rows)
    # Run-level badges use the first experiment's settings. Extra experiments
    # still get their own plate sections from `Analysis_graph`.
    if blocks:
        first = blocks[0]
        for label, key in _INFO_METADATA_KEYS.items():
            if label in first and first[label] not in (None, ""):
                metadata[key] = _jsonable(first[label])
    return metadata


def _info_blocks(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    starts = [i for i, value in enumerate(rows[0]) if value == "Experiment settings"]
    if not starts:
        # Single two-column table with labels in column 0.
        starts = [0]
    width = starts[1] - starts[0] if len(starts) > 1 else min(2, len(rows[0]))
    blocks: list[dict[str, Any]] = []
    for start in starts:
        block: dict[str, Any] = {}
        for row in rows:
            label = _cell_str(_cell_at(row, start))
            if not label:
                continue
            block[label] = _cell_at(row, start + 1) if width > 1 else None
        blocks.append(block)
    return blocks


def _series_pairs_from_headers(
    headers: tuple[Any, ...],
) -> tuple[str, list[tuple[str, int, int]]] | None:
    """Walk (x, y) column pairs and keep the ones we know how to plot."""
    names = [_cell_str(h) for h in headers]
    pairs: list[tuple[str, int, int]] = []
    flavor: str | None = None
    hydro_seen = 0
    i = 0
    while i + 1 < len(names):
        matched = _match_header_pair(names[i], names[i + 1], hydro_seen)
        if matched is None:
            i += 1
            continue
        series_id, pair_flavor = matched
        if pair_flavor == FLAVOR_SIZING and names[i] == "hydrodynamic_diameter":
            hydro_seen += 1
        if flavor is None:
            flavor = pair_flavor
        if pair_flavor == flavor:
            pairs.append((series_id, i, i + 1))
        i += 2
    if flavor is None or not pairs:
        return None
    return flavor, pairs


def _match_header_pair(
    x_name: str | None, y_name: str | None, hydro_seen: int
) -> tuple[str, str] | None:
    if x_name == "hydrodynamic_diameter" and y_name == "amplitude":
        series_id = SERIES_INTENSITY if hydro_seen == 0 else SERIES_MASS
        return series_id, FLAVOR_SIZING
    for x_header, y_header, pair_flavor, series_id in _SERIES_PAIR_KEYS:
        if x_name == x_header and y_name == y_header:
            return series_id, pair_flavor
    return None


def _find_row(rows: list[tuple[Any, ...]], label: str) -> int | None:
    for i, row in enumerate(rows):
        if row and row[0] == label:
            return i
    return None


def _find_series_row(rows: list[tuple[Any, ...]], after: int) -> int | None:
    for i in range(after + 1, len(rows)):
        row = rows[i]
        if not row:
            continue
        first = _cell_str(row[0])
        if first in _SERIES_ROW_FIRST_CELLS:
            return i
    return None


def _row_slice(row: tuple[Any, ...], start: int, width: int) -> tuple[Any, ...]:
    end = start + width
    padded = list(row[start:end])
    if len(padded) < width:
        padded.extend([None] * (width - len(padded)))
    return tuple(padded)


def _cell_at(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _num(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        try:
            number = float(str(value).strip())
        except (TypeError, ValueError):
            return None
    if math.isnan(number) or math.isinf(number):
        return None
    return number


def _table_value(key: str, value: float | None) -> float | None:
    if value is None:
        return None
    if key in _TM_VALUE_KEYS and value == 0:
        return None
    return value


def _jsonable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = _num(value)
        return number
    if isinstance(value, str):
        return value
    return str(value)
