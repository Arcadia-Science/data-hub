"""Synthetic Aunty workbooks for parser tests and local seed fixtures."""

from __future__ import annotations
import math
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook


@dataclass
class SyntheticWell:
    well: str
    sample: str | None = None
    values: dict[str, float | None] = field(default_factory=dict)
    series: dict[str, list[tuple[float, float]]] = field(default_factory=dict)


@dataclass
class SyntheticExperiment:
    file_name: str
    flavor: str
    analysis_mode: str | None = None
    info: dict[str, object] = field(default_factory=dict)
    wells: list[SyntheticWell] = field(default_factory=list)


def write_aunty_workbook(
    path: Path,
    experiments: list[SyntheticExperiment],
    *,
    include_graph: bool = True,
) -> None:
    """Write an Aunty-shaped `.xlsx` covering the graph, table, and info sheets."""
    if not experiments:
        raise ValueError("Need at least one experiment")
    flavor = experiments[0].flavor
    if any(exp.flavor != flavor for exp in experiments):
        raise ValueError("A workbook must use one flavor so column blocks line up")

    wb = Workbook()
    first = wb.active
    assert first is not None
    if include_graph:
        first.title = "Analysis_graph"
        _write_graph(first, experiments, flavor)
        results = wb.create_sheet("Results")
    else:
        first.title = "Results"
        results = first
    _write_results(results, experiments, flavor)

    table = wb.create_sheet("Analysis_table")
    _write_table(table, experiments, flavor)

    info = wb.create_sheet("Experiment info")
    _write_info(info, experiments, flavor)

    ident = wb.create_sheet("ExperimentID")
    ident["A1"] = "00000000-0000-4000-8000-000000000001"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def thermal_experiment(
    file_name: str,
    wells: list[str],
    *,
    n_points: int = 8,
    tm1: float = 64.6,
    include_series: set[str] | None = None,
) -> SyntheticExperiment:
    """A small BCM thermal-ramp experiment with a fluorescence peak near Tm1."""
    experiment = SyntheticExperiment(
        file_name=file_name,
        flavor="thermal_ramp",
        analysis_mode="BCM",
        info={
            "Start temp (°C)": 25,
            "End temp (°C)": 95,
            "Rate (°C/min)": 1,
            "Analysis version": "v2.0.1",
            "Experiment name": file_name,
        },
    )
    temps = [25.0 + i * (70.0 / max(n_points - 1, 1)) for i in range(n_points)]
    for i, well in enumerate(wells):
        well_tm = tm1 + i * 0.4
        fluorescence = [(t, 330.0 + 40.0 * _peak(t, well_tm, 8.0)) for t in temps]
        differential = [(t, _deriv(t, well_tm, 8.0)) for t in temps[:-1]]
        diameter = [(t, 80.0 + (t - 25.0) * 2.0) for t in temps]
        sls = [(t, 1.0e6 + (t - 25.0) * 1.0e5) for t in temps]
        series = {
            "fluorescence": fluorescence,
            "differential": differential,
            "z_average_diameter": diameter,
            "sls": sls,
        }
        if include_series is not None:
            series = {key: points for key, points in series.items() if key in include_series}
        experiment.wells.append(
            SyntheticWell(
                well=well,
                sample=well,
                values={
                    "Tm1 (°C)": well_tm,
                    "Tm2 (°C)": 0,
                    "Tagg (°C)": well_tm - 4.0,
                    "Tonset (°C)": well_tm - 8.0,
                    "Tsize (°C)": 39.0 + i,
                },
                series=series,
            )
        )
    return experiment


def sizing_experiment(
    file_name: str,
    wells: list[str],
    *,
    n_points: int = 12,
    z_avg: float = 75.0,
) -> SyntheticExperiment:
    """A sizing experiment with correlation plus intensity/mass distributions."""
    experiment = SyntheticExperiment(
        file_name=file_name,
        flavor="sizing",
        info={
            "Temperature (°C)": 25,
            "Analysis version": "v2.0.1",
            "Experiment name": file_name,
        },
    )
    times = [4e-7 * (1.5**i) for i in range(n_points)]
    diameters = [0.5 * (1.4**i) for i in range(n_points)]
    for i, well in enumerate(wells):
        well_z = z_avg + i * 5.0
        intensity_peak = 12.0 + i
        mass_peak = 2.0 + i * 0.2
        experiment.wells.append(
            SyntheticWell(
                well=well,
                sample=well,
                values={
                    "Z-avg diameter (nm)": well_z,
                    "PDI": 0.3 + i * 0.01,
                    "Pk 1 diameter (nm)": intensity_peak,
                    "Pk 1 intensity (%)": 60.0,
                    "Pk 1 mass (%)": 1.0,
                },
                series={
                    "correlation": [(t, 0.85 * (0.7 ** (j / 2))) for j, t in enumerate(times)],
                    "intensity": [(d, 0.2 * _peak(d, intensity_peak, 30.0)) for d in diameters],
                    "mass": [(d, 0.2 * _peak(d, mass_peak, 2.0)) for d in diameters],
                },
            )
        )
    return experiment


def isothermal_experiment(
    file_name: str,
    wells: list[str],
    *,
    n_points: int = 8,
    with_sls: bool = False,
    k1: float = 0.0034,
) -> SyntheticExperiment:
    """An isothermal hold with time-series fluorescence, optionally SLS."""
    experiment = SyntheticExperiment(
        file_name=file_name,
        flavor="isothermal",
        analysis_mode="Peak height",
        info={
            "Temperature (°C)": 25,
            "Analysis version": "v2.0.1",
            "Experiment name": file_name,
        },
    )
    times = [float(i * 60) for i in range(n_points)]
    for i, well in enumerate(wells):
        well_k1 = k1 + i * 0.0002
        series: dict[str, list[tuple[float, float]]] = {
            "fluorescence": [(t, 22_000.0 - t * (0.4 + i * 0.05)) for t in times],
        }
        values: dict[str, float | None] = {
            "fluorescence k₁ (s⁻¹)": well_k1,
            "fluorescence k₂ (s⁻¹)": None,
            "fluorescence R²": 0.95,
        }
        if with_sls:
            series["sls"] = [(t, 3.5e6 + t * 200.0) for t in times]
            values["SLS k₁ (s⁻¹)"] = well_k1 * 0.4
        experiment.wells.append(
            SyntheticWell(well=well, sample=well, values=values, series=series)
        )
    return experiment


def _peak(x: float, center: float, width: float) -> float:
    return math.exp(-((x - center) ** 2) / (2 * width * width))


def _deriv(x: float, center: float, width: float) -> float:
    return -((x - center) / (width * width)) * _peak(x, center, width)


_GRAPH_SERIES: dict[str, list[tuple[str, tuple[str, str]]]] = {
    "thermal_ramp": [
        ("fluorescence", ("temperature_fluorescence", "fluorescence")),
        ("differential", ("temperature_differential", "differential")),
        ("z_average_diameter", ("temperature_DLS", "Z-average diameter")),
        ("sls", ("temperature_DLS", "SLS")),
    ],
    "sizing": [
        ("correlation", ("time", "amplitude")),
        ("intensity", ("hydrodynamic_diameter", "amplitude")),
        ("mass", ("hydrodynamic_diameter", "amplitude")),
    ],
    "isothermal": [
        ("fluorescence", ("time_fluorescence", "fluorescence")),
        ("sls", ("time_DLS", "SLS")),
    ],
}


def _write_graph(ws, experiments: list[SyntheticExperiment], flavor: str) -> None:
    present = set(experiments[0].wells[0].series)
    spec = [(key, headers) for key, headers in _GRAPH_SERIES[flavor] if key in present]
    if not spec:
        raise ValueError(f"No graph series to write for flavor {flavor}")
    series_keys = [key for key, _headers in spec]
    header_labels = [label for _key, headers in spec for label in headers]
    width = len(header_labels)
    has_mode = flavor in {"thermal_ramp", "isothermal"}

    # Flatten to (experiment, well) in the same column-major caller order.
    pairs: list[tuple[SyntheticExperiment, SyntheticWell]] = []
    for exp in experiments:
        for well in exp.wells:
            pairs.append((exp, well))

    # Header stack: File name / optional Analysis mode / Well / Sample / blank / series.
    row = 1
    for col_i, (exp, _well) in enumerate(pairs):
        start = col_i * width + 1
        ws.cell(row, start, "File name")
        ws.cell(row, start + 1, exp.file_name)
    row += 1
    if has_mode:
        for col_i, (exp, _well) in enumerate(pairs):
            start = col_i * width + 1
            ws.cell(row, start, "Analysis mode")
            ws.cell(row, start + 1, exp.analysis_mode)
        row += 1
    for col_i, (_exp, well) in enumerate(pairs):
        start = col_i * width + 1
        ws.cell(row, start, "Well")
        ws.cell(row, start + 1, well.well)
    row += 1
    for col_i, (_exp, well) in enumerate(pairs):
        start = col_i * width + 1
        ws.cell(row, start, "Sample")
        ws.cell(row, start + 1, well.sample)
    row += 1
    # Blank row, then series names.
    row += 1
    series_row = row
    for col_i, _pair in enumerate(pairs):
        start = col_i * width + 1
        for offset, label in enumerate(header_labels):
            ws.cell(series_row, start + offset, label)

    data_start = series_row + 1
    for col_i, (_exp, well) in enumerate(pairs):
        start = col_i * width + 1
        for series_i, series_id in enumerate(series_keys):
            points = well.series.get(series_id, [])
            x_col = start + series_i * 2
            y_col = start + series_i * 2 + 1
            for point_i, (x, y) in enumerate(points):
                ws.cell(data_start + point_i, x_col, x)
                ws.cell(data_start + point_i, y_col, y)


def _write_table(ws, experiments: list[SyntheticExperiment], flavor: str) -> None:
    if flavor == "thermal_ramp":
        headers = [
            "Index",
            "File name",
            "Analysis mode",
            "Well",
            "Sample",
            "Tm1 (°C)",
            "Tm2 (°C)",
            "Tm3 (°C)",
            "Tagg (°C)",
            "Tonset (°C)",
            "Tsize (°C)",
        ]
        value_keys = [
            "Tm1 (°C)",
            "Tm2 (°C)",
            "Tm3 (°C)",
            "Tagg (°C)",
            "Tonset (°C)",
            "Tsize (°C)",
        ]
    elif flavor == "sizing":
        headers = [
            "Index",
            "File name",
            "Well",
            "Sample",
            "Z-avg diameter (nm)",
            "PDI",
            "Pk 1 diameter (nm)",
            "Pk 1 intensity (%)",
            "Pk 1 mass (%)",
        ]
        value_keys = [
            "Z-avg diameter (nm)",
            "PDI",
            "Pk 1 diameter (nm)",
            "Pk 1 intensity (%)",
            "Pk 1 mass (%)",
        ]
    else:
        headers = [
            "Index",
            "File name",
            "Analysis mode",
            "Well",
            "Sample",
            "fluorescence k₁ (s⁻¹)",
            "fluorescence k₂ (s⁻¹)",
            "fluorescence R²",
            "SLS k₁ (s⁻¹)",
        ]
        value_keys = [
            "fluorescence k₁ (s⁻¹)",
            "fluorescence k₂ (s⁻¹)",
            "fluorescence R²",
            "SLS k₁ (s⁻¹)",
        ]

    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)

    row_i = 2
    index = 1
    for exp in experiments:
        for well in exp.wells:
            ws.cell(row_i, 1, index)
            ws.cell(row_i, 2, exp.file_name)
            col = 3
            if flavor in {"thermal_ramp", "isothermal"}:
                ws.cell(row_i, col, exp.analysis_mode)
                col += 1
            ws.cell(row_i, col, well.well)
            ws.cell(row_i, col + 1, well.sample)
            for offset, key in enumerate(value_keys):
                ws.cell(row_i, col + 2 + offset, well.values.get(key))
            row_i += 1
            index += 1


def _write_results(ws, experiments: list[SyntheticExperiment], flavor: str) -> None:
    if flavor == "thermal_ramp":
        headers = [
            "File name",
            "Analysis mode",
            "Selected",
            "Well",
            "Sample",
            "ID",
            "Tm1 (°C)",
        ]
    elif flavor == "isothermal":
        headers = [
            "File name",
            "Analysis mode",
            "Selected",
            "Well",
            "Sample",
            "ID",
            "k₁ (s⁻¹) fluorescence",
        ]
    else:
        headers = ["File name", "Selected", "Well", "Sample", "ID", "Z-avg diameter (nm)"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    row_i = 2
    index = 1
    for exp in experiments:
        for well in exp.wells:
            ws.cell(row_i, 1, exp.file_name)
            col = 2
            if flavor in {"thermal_ramp", "isothermal"}:
                ws.cell(row_i, col, exp.analysis_mode)
                col += 1
            ws.cell(row_i, col, "x")
            ws.cell(row_i, col + 1, well.well)
            ws.cell(row_i, col + 2, well.sample)
            ws.cell(row_i, col + 3, index)
            if flavor == "thermal_ramp":
                ws.cell(row_i, col + 4, well.values.get("Tm1 (°C)"))
            elif flavor == "isothermal":
                ws.cell(row_i, col + 4, well.values.get("fluorescence k₁ (s⁻¹)"))
            else:
                ws.cell(row_i, col + 4, well.values.get("Z-avg diameter (nm)"))
            row_i += 1
            index += 1


def _write_info(ws, experiments: list[SyntheticExperiment], flavor: str) -> None:
    for block_i, exp in enumerate(experiments):
        start = block_i * 2 + 1
        ws.cell(1, start, "Experiment settings")
        ws.cell(3, start, "File name")
        ws.cell(3, start + 1, exp.file_name)
        if flavor in {"thermal_ramp", "isothermal"}:
            ws.cell(4, start, "Analysis mode")
            ws.cell(4, start + 1, exp.analysis_mode)
        row = 5
        for label, value in exp.info.items():
            ws.cell(row, start, label)
            ws.cell(row, start + 1, value)
            row += 1


def column_major_96() -> list[str]:
    """Aunty walks the plate column-major: A1, B1, … H1, A2, …"""
    return [f"{row}{col}" for col in range(1, 13) for row in "ABCDEFGH"]


if __name__ == "__main__":
    fixtures = Path(__file__).resolve().parents[1] / "fixtures"
    wells = column_major_96()
    write_aunty_workbook(
        fixtures / "aunty_thermal_ramp.xlsx",
        [thermal_experiment("Thermal ramp seed T0900", wells, n_points=16)],
    )
    write_aunty_workbook(
        fixtures / "aunty_sizing.xlsx",
        [sizing_experiment("Sizing seed T1000", wells, n_points=16)],
    )
    write_aunty_workbook(
        fixtures / "aunty_isothermal.xlsx",
        [
            isothermal_experiment(
                "Blue Isothermal seed T1016",
                wells[:12],
                with_sls=True,
            )
        ],
    )
    write_aunty_workbook(
        fixtures / "aunty_table_only.xlsx",
        [thermal_experiment("Thermal ramp seed T1627", wells[:24])],
        include_graph=False,
    )
